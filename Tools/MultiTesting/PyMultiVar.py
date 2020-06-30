#install openpyxl (pip install )
#open cmd with dir to this folder
#run python

#-----------IMPORTS---------------------
import os
import time 
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border,colors
from openpyxl.cell import Cell

#-----------CONSTANTS-------------------
SURV_NAME = "VFU"
VAR_1 = "jumpDist"
MIN_1 = 0x0000
MAX_1 = 0x6000
DELTA_1 = 0x100
VAR_2 = "callAmount"
MIN_2 = 0x40
MAX_2 = 0xC0
DELTA_2 = 0x1

FILE_PATH = os.getcwd()+"\\"+SURV_NAME
EXCEL_FILENAME = "multi_data.xlsx"
EXCEL_FOLDER = "excel_data_folder"
GREEN_FILL = PatternFill(start_color='FF00FF00',end_color = 'FF00FF00',fill_type='solid')
#----------------GLOBALS----------------
cur_excel_idx = 0
workbook = None
#-----------UTIL------------------------

# Print iterations progress (from StackOverFlow)
def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print('\r%s |%s| %s%% %s' % (prefix, bar, percent, suffix), end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()        
        

# returns list of survivors' names from folder + my survivor
def get_surv_names():
    files_in_dir = os.listdir(os.getcwd()+"\\survivors\\")
    survivors = set()
    for name in files_in_dir:
        if name[-1] == '1' or name[-1] == '2': # remove number of survivor from its name
            survivors.add(name[:-1])
        else: # if lone survivor
            survivors.add(name)
    res = list(survivors)
    if SURV_NAME not in res:
        res = [SURV_NAME] + res
    return res
    

# returns the name of the x'th excel file
def excel_filename(x):
    return EXCEL_FILENAME.split(".")[0]+str(x)+"."+EXCEL_FILENAME.split(".")[1]
    
    
# preparing the excel file - writing the headers for the table and creating the different sheets
def prepare_excel_files(names):
    workbook = Workbook()
    # changing the sheets names
    workbook.remove(workbook[workbook.sheetnames[0]])
    for name in names:
        workbook.create_sheet(name)
    
    # filling each sheet with the table headers
    for name in names:
        cur = workbook[name]
        cur["A1"] = VAR_1+"\\"+VAR_2
        # col-header
        idx = 2
        for v1 in range(MIN_1, MAX_1+1, DELTA_1):
            cur[f"A{idx}"] = hex(v1)[2:]
            idx+=1
        # row-header    
        idx = 2
        for v2 in range(MIN_2,MAX_2+1,DELTA_2):
            cur[f"{get_column_letter(idx)}1"] = hex(v2)[2:]
            idx += 1
            
    #saving the file
    workbook.save(filename = os.getcwd()+"\\"+EXCEL_FOLDER+"\\"+excel_filename(0))


# loads the excel table into global var "workbook"
def load_excel():
    global workbook,cur_excel_idx
    workbook = load_workbook(filename=os.getcwd()+"\\"+EXCEL_FOLDER+"\\"+excel_filename(cur_excel_idx))

# read the data from the csv file
# returns a list where each element is [nameOfSurv, scoreOfSurv]
def read_csv():
    f = open("scores.csv","r+")
    lines = f.readlines()
    lines = [i.replace("\n","") for i in lines]
    idx = 1
    data = []
    while lines[idx] != "":
        data.append(lines[idx].split(','))
        data[-1][1] = float(data[-1][1])
        idx+=1
        
    f.close()
    return data
    
# returns list of round winners
def get_best_surv(data):
    winners = []
    max_score = -1
    for name,score in data:
        if score==max_score:
            winners.append(name)
        elif score>max_score:
            max_score = score
            winners = [name]
    return winners
    
    
#updates the workbook variable according the the results from csv
def update_excel(data,idx1,idx2):
    global workbook
    winners = get_best_surv(data)
    cell_idx = f"{get_column_letter(idx2+2)}{idx1+2}"
    for name,score in data:
        workbook[name][cell_idx] = score
        if name in winners:
            workbook[name][cell_idx].fill = GREEN_FILL
    
 
# save the workbook into an excel file
def save_excel():
    global workbook,cur_excel_idx
    workbook.save(filename = os.getcwd()+"\\"+EXCEL_FOLDER+"\\"+excel_filename(cur_excel_idx+1))
    cur_excel_idx +=1


#------------CODE-----------------------

#prepare excel file
os.system("mkdir "+ EXCEL_FOLDER)
prepare_excel_files(get_surv_names())
# get ready for autotuning - going through the code files and finding the lines of the defined variables
lines = [None,None] 
var1 = [-1,-1]
var2 = [-1,-1]
for i in [1,2]:
    surv = open(FILE_PATH+f"{i}.asm","r+")
    lines[i-1] = surv.readlines()
    #print(lines[i-1])
    for idx in range(len(lines[i-1])):
        if lines[i-1][idx].startswith(f"%define {VAR_1}"):
            var1[i-1] = idx
        elif lines[i-1][idx].startswith(f"%define {VAR_2}"):
            var2[i-1] = idx
            
            
# autotune
idx1 = 0
for v1 in range(MIN_1, MAX_1+1, DELTA_1):
    printProgressBar(0, MAX_2+1-MIN_2, prefix = f"{VAR_1} = {hex(v1)}/{hex(MAX_1)}:{(len(str(MAX_2))-len(str(v1)))*' '}", suffix = 'Complete', length = 50)
    # loads excel table
    load_excel()
    idx2 = 0
    for v2 in range(MIN_2,MAX_2+1,DELTA_2):
        # write and compile with new values
        for i in [1,2]:
            temp = open(os.getcwd()+f"\\temp{i}.asm","w+")
            lines[i-1][var1[i-1]] = f"%define {VAR_1} {v1}\n"
            lines[i-1][var2[i-1]] = f"%define {VAR_2} {v2}\n"
            temp.writelines(lines[i-1])
            temp.close()
            os.system(f"nasm temp{i}.asm -o ./survivors/{SURV_NAME}{i}")
            
        #run cgx
        os.system(f"multi.jar")
        
        #wait for engine to finish
        while not os.path.isfile('./scores.csv'):
            time.sleep(0.1)
        
        #collect results and update the table accordingly 
        update_excel(read_csv(),idx1,idx2)
        
        #erase csv file
        os.remove(os.getcwd()+f"\\scores.csv")
        printProgressBar(v2+1-MIN_2, MAX_2+1-MIN_2, prefix = f"{VAR_1} = {hex(v1)}/{hex(MAX_1)}:{(len(str(MAX_2))-len(str(v1)))*' '}", suffix = 'Complete', length = 50)
        idx2+=1
    
    #save table to excel file
    save_excel()
    idx1+=1
    
    
        
#delete temp files
for i in [1,2]:
    os.remove(os.getcwd()+f"\\temp{i}.asm")
    os.remove(os.getcwd()+f"\\survivors\\{SURV_NAME}{i}")

print("Done!")